import random
import time
import os
from datetime import datetime

from playwright.sync_api import sync_playwright
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


class BaseScraper:

    def __init__(self, proxy=None):

        self.browser = None
        self.page = None
        self.playwright = None
        self.proxy = proxy

        self.start_time = None
        self.end_time = None

        today = datetime.now().strftime("%Y_%m_%d")

        os.makedirs("job_reports", exist_ok=True)

        self.excel_file = f"job_reports/jobs_{today}.xlsx"

    # --------------------------------
    # Start timer
    # --------------------------------
    def start_timer(self):
        self.start_time = time.time()

    def stop_timer(self):
        self.end_time = time.time()

    def get_execution_time(self):

        if not self.start_time or not self.end_time:
            return "Unknown"

        seconds = int(self.end_time - self.start_time)

        minutes = seconds // 60
        seconds = seconds % 60

        return f"{minutes}m {seconds}s"

    # --------------------------------
    # Random Delay (optimized)
    # --------------------------------
    def random_delay(self, min_s=1.2, max_s=3.2):
        time.sleep(random.uniform(min_s, max_s))

    # --------------------------------
    # Random User Agent
    # --------------------------------
    def get_random_user_agent(self):

        user_agents = [

            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/120.0 Safari/537.36",
            "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 Chrome/119.0 Safari/537.36",
            "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 Chrome/118.0 Safari/537.36"

        ]

        return random.choice(user_agents)

    # --------------------------------
    # Start Browser
    # --------------------------------
    def start_browser(self):

        print("Launching browser...")

        self.playwright = sync_playwright().start()

        launch_args = [
            "--disable-blink-features=AutomationControlled",
            "--disable-infobars",
            "--no-sandbox",
            "--disable-dev-shm-usage"
        ]

        browser_args = {

            "channel": "chrome",
            "headless": False,
            "slow_mo": random.randint(5, 12),
            "args": launch_args

        }

        if self.proxy:
            browser_args["proxy"] = {"server": self.proxy}

        self.browser = self.playwright.chromium.launch(**browser_args)

        context = self.browser.new_context(

            user_agent=self.get_random_user_agent(),
            viewport={"width": 1366, "height": 768},
            locale="en-US",
            timezone_id="Asia/Kolkata"

        )

        context.add_init_script(
            """Object.defineProperty(navigator, 'webdriver', {get: () => undefined})"""
        )

        self.page = context.new_page()

        print("Browser started")

    # --------------------------------
    # Safe navigation
    # --------------------------------
    def safe_goto(self, url, retries=3):

        for attempt in range(retries):

            try:

                print("Navigating:", url)

                self.page.goto(url, timeout=60000, wait_until="domcontentloaded")

                self.random_delay()

                return True

            except Exception as e:

                print("Navigation failed:", e)

                if attempt < retries - 1:

                    print("Retrying navigation...")
                    self.random_delay(2, 4)

        return False

    # --------------------------------
    # Scroll page (optimized)
    # --------------------------------
    def scroll_page(self, scroll_count=4):

        for i in range(scroll_count):

            scroll_value = random.randint(1500, 3500)

            print("Scrolling:", i + 1)

            try:
                self.page.mouse.wheel(0, scroll_value)
            except:
                self.page.evaluate(f"window.scrollBy(0,{scroll_value})")

            self.random_delay(1, 2)

    # --------------------------------
    # Save Excel (same as yours)
    # --------------------------------
    def save_jobs_to_excel(self, jobs):

        if not jobs:
            return

        headers = [
            "Platform",
            "Title",
            "Company",
            "Location",
            "Search Location",
            "Level",
            "Posted Time",
            "Job Link",
            "Scraped At"
        ]

        if not os.path.exists(self.excel_file):

            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Jobs"

            sheet.append(headers)

        else:

            workbook = load_workbook(self.excel_file)
            sheet = workbook.active

        for job in jobs:

            sheet.append([

                job.get("platform"),
                job.get("title"),
                job.get("company"),
                job.get("location"),
                job.get("search_location"),
                job.get("search_level"),
                job.get("posted_time"),
                f'=HYPERLINK("{job.get("job_link")}", "Open Job")',
                datetime.now().strftime("%Y-%m-%d %H:%M")

            ])

        workbook.save(self.excel_file)

        print("Excel saved:", self.excel_file)

    # --------------------------------
    # Close browser
    # --------------------------------
    def close_browser(self):

        if self.browser:
            self.browser.close()

        if self.playwright:
            self.playwright.stop()

        print("Browser closed")
